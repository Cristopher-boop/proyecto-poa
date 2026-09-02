import os
import re
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone

from apps.presupuestos.models import Gestion, Partida, PresupuestoArea
from apps.organizacional.models import Area, Seccion
from apps.planificacion.models import Operacion
from apps.memorias.models import MemoriaCalculo, DetallePresupuestoMemoria
from apps.memorias.utils import recalcular_saldos_memoria

def text(value) -> str:
    return str(value or '').strip()

def key(value) -> str:
    normalized = unicodedata.normalize('NFKD', text(value)).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'[^A-Z0-9]+', ' ', normalized.upper()).strip()

def decimal(value, default=Decimal('0')) -> Decimal:
    if value is None or text(value) == '':
        return default
    try:
        return Decimal(str(value).replace(',', '.')).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        return default

def find_memory_header(sheet):
    columns = {'numero': 0, 'descripcion': 1, 'cantidad': 3, 'unidad': 4, 'precio': 5, 'total': 6, 'justificacion': 7}
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        labels = [key(value) for value in row]
        has_desc = any('DESCRIPCI' in label for label in labels)
        has_other = any(any(k in label for k in ['CANTIDAD', 'PRECIO', 'UNIDAD', 'TOTAL', 'P U', 'UNITARIO']) for label in labels)
        if not (has_desc and has_other):
            continue
        price_column_found = False
        for index, label in enumerate(labels):
            if label in {'N', 'NO', 'NRO', 'NUMERO', 'N ', 'ITEM'}:
                columns['numero'] = index
            elif 'DESCRIPCI' in label:
                columns['descripcion'] = index
            elif 'CANTIDAD' in label:
                columns['cantidad'] = index
            elif 'UNIDAD' in label:
                columns['unidad'] = index
            elif 'PRECIO' in label or 'UNITARIO' in label or 'P U' in label or 'P.U' in label:
                if price_column_found:
                    columns['total'] = index
                else:
                    columns['precio'] = index
                    price_column_found = True
            elif label == 'TOTAL':
                columns['total'] = index
            elif 'JUSTIFICACI' in label:
                columns['justificacion'] = index
        return row_number, columns
    return None, None

AREA_MAP = {
    'PL': 'PL',
    'UT': 'UT',
    'AI': 'AI',
    'AJ': 'AJ',
    'CIAC': 'CIAC',
    'OD': 'OD',
    'IF': 'IF',
    'GAA': 'GAA',
    'GC': 'GC',
    'CG': 'GC',
    'GO': 'GO',
    'AE': 'AE',
    'SMS': 'SMS',
    'EA': 'EA',
    'LP': 'LP',
    'CBBA': 'CB',
    'CB': 'CB',
    'SCZ': 'SC',
    'SC': 'SC',
    'CBJ': 'CIJ',
    'CIJ': 'CIJ',
    'GYA': 'GYA',
    'RIB': 'RIB',
    'TDD': 'TDD',
}

class Command(BaseCommand):
    help = 'Carga las Memorias de Cálculo oficiales de la gestión 2026 desde el Excel CONSOLIDADO POA 2026 OFICIAL, descartando las partidas con total 0,00'

    def add_arguments(self, parser):
        parser.add_argument('--consolidado', help='Ruta al archivo Excel oficial 2026.')

    def handle(self, *args, **options):
        project_dir = Path(__file__).resolve().parents[4]
        default_excel = project_dir / '1. CONSOLIDADO GENERAL POA 2026 OFICIAL pal mauro.xlsx'
        if not default_excel.is_file():
            default_excel = project_dir.parent / '1. CONSOLIDADO GENERAL POA 2026 OFICIAL pal mauro.xlsx'

        excel_path = Path(options.get('consolidado') or os.getenv('POA_2026_CONSOLIDADO', default_excel)).expanduser().resolve()

        if not excel_path.is_file():
            self.stdout.write(self.style.ERROR(f"No se encontró el archivo Excel oficial 2026 en: {excel_path}"))
            return

        self.stdout.write(self.style.WARNING("=" * 60))
        self.stdout.write(self.style.WARNING(" CARGA OFICIAL POA GESTIÓN 2026 (MEMORIAS DE CÁLCULO)"))
        self.stdout.write(self.style.WARNING("=" * 60))
        self.stdout.write(f"Archivo: {excel_path.name}")

        with transaction.atomic():
            # Asegurar Gestión 2026
            gestion, _ = Gestion.objects.get_or_create(
                anio=2026,
                defaults={'estado': Gestion.EstadoGestion.FORMULACION}
            )

            # Limpiar memorias existentes de 2026 y resetear presupuestos de 2026
            MemoriaCalculo.objects.filter(gestion=gestion).delete()
            PresupuestoArea.objects.filter(gestion=gestion).update(
                monto_inicial=Decimal('0.00'),
                monto_actual=Decimal('0.00')
            )

            self.stdout.write("Abriendo libro Excel oficial 2026...")
            wb = openpyxl.load_workbook(excel_path, data_only=True)

            sheets = wb.sheetnames[13:]  # Hojas de memorias individuales
            self.stdout.write(f"Analizando {len(sheets)} hojas de memorias de cálculo...")

            count_memorias = 0
            count_detalles = 0
            count_omitidas_cero = 0
            total_presupuesto_2026 = Decimal('0.00')
            area_counters = {}
            area_totals = {}

            for sname in sheets:
                ws = wb[sname]
                sname_clean = sname.strip()

                # Extraer partida (ej. '22110PL' -> '22110')
                partida_match = re.search(r'\d{5}', sname_clean)
                partida_code = partida_match.group(0) if partida_match else None
                if not partida_code:
                    continue

                # Extraer sigla de unidad
                sigla_match = re.search(r'[A-Za-z]+', sname_clean)
                sigla_raw = sigla_match.group(0).upper() if sigla_match else ''
                area_sigla = AREA_MAP.get(sigla_raw, sigla_raw)

                # Identificar cabecera
                header_r, cols = find_memory_header(ws)
                if not header_r:
                    self.stdout.write(self.style.WARNING(f"  [!] Cabecera no encontrada en {sname}, omitiendo..."))
                    continue

                # Extraer ítems de la hoja y el valor en TOTAL PARTIDA
                items = []
                source_total_partida = None

                for row in ws.iter_rows(min_row=header_r + 1, values_only=True):
                    values = list(row)
                    desc = text(values[cols['descripcion']]) if len(values) > cols['descripcion'] else ''
                    num = text(values[cols['numero']]) if len(values) > cols['numero'] else ''
                    row_str = ' '.join([key(v) for v in values if v is not None])

                    if 'TOTAL' in key(num) or 'TOTAL' in key(desc) or 'TOTAL PARTIDA' in row_str:
                        candidate_val = None
                        if len(values) > cols['total'] and values[cols['total']] is not None:
                            candidate_val = values[cols['total']]
                        else:
                            for v in reversed(values):
                                if isinstance(v, (int, float, Decimal)) or (isinstance(v, str) and re.match(r'^-?\d+(?:[\.,]\d+)?$', v.strip())):
                                    candidate_val = v
                                    break
                        source_total_partida = decimal(candidate_val)
                        break

                    if key(num) in {'ELABORADO POR', 'APROBADO POR', 'REVISADO POR'} or key(desc) in {'ELABORADO POR', 'APROBADO POR', 'REVISADO POR'}:
                        break
                    if not desc:
                        continue

                    cant = decimal(values[cols['cantidad']] if len(values) > cols['cantidad'] else None, Decimal('1.00'))
                    unit = text(values[cols['unidad']] if len(values) > cols['unidad'] else '') or 'UNIDAD'
                    price = decimal(values[cols['precio']] if len(values) > cols['precio'] else None)
                    tot = decimal(values[cols['total']] if len(values) > cols['total'] else None)
                    if tot == Decimal('0.00') and (cant * price) > 0:
                        tot = (cant * price).quantize(Decimal('0.01'))

                    base_tot = cant * price
                    factor = (tot / base_tot).quantize(Decimal('0.0001')) if base_tot else Decimal('1.0000')
                    just = text(values[cols['justificacion']] if len(values) > cols['justificacion'] else '')

                    items.append({
                        'descripcion': desc,
                        'cantidad': cant,
                        'unidad_medida': unit[:50],
                        'precio_unitario': price,
                        'factor_calculo': factor,
                        'total_origen': tot,
                        'justificacion': just
                    })

                items_sum = sum((it['total_origen'] for it in items), Decimal('0.00')).quantize(Decimal('0.01'))

                # REGLA CRÍTICA: Si la celda de TOTAL PARTIDA es 0,00 o la suma es 0,00, se DESCARTA / ELIMINA
                if (source_total_partida is not None and source_total_partida <= Decimal('0.00')) or (source_total_partida is None and items_sum <= Decimal('0.00')) or not items:
                    count_omitidas_cero += 1
                    continue

                sheet_total = source_total_partida if source_total_partida is not None else items_sum

                # Localizar Área y Sección
                area = Area.objects.filter(codigo__iendswith=f"-{area_sigla}").first()
                if not area:
                    area = Area.objects.filter(codigo__icontains=area_sigla).first()
                if not area:
                    area = Area.objects.first()

                seccion = Seccion.objects.filter(area=area).first()
                if not seccion:
                    seccion, _ = Seccion.objects.get_or_create(area=area, defaults={'nombre': area.nombre})

                # Localizar Operación 2026 para esta área
                operacion = Operacion.objects.filter(area=area, accion_corto_plazo__gestion=gestion).first()
                if not operacion:
                    operacion = Operacion.objects.filter(area=area).first()

                # Localizar o crear Partida de Egreso
                partida, _ = Partida.objects.get_or_create(
                    codigo=partida_code,
                    clase=Partida.ClasePartida.EGRESO,
                    defaults={'nombre': f"Partida {partida_code}"}
                )

                # Generar código correlativo de memoria
                correlativo = area_counters.get(area_sigla, 0) + 1
                area_counters[area_sigla] = correlativo
                memoria_code = f"MEM-{area_sigla}-2026-{correlativo:03d}"

                # Justificación general
                justificacion_memoria = next((it['justificacion'] for it in items if it['justificacion']), '')
                if not justificacion_memoria:
                    justificacion_memoria = f"Memoria de Cálculo de la partida {partida_code} ({sname_clean}) para {area.nombre}."

                # Crear Memoria
                memoria = MemoriaCalculo.objects.create(
                    codigo=memoria_code,
                    gestion=gestion,
                    seccion=seccion,
                    operacion=operacion,
                    justificacion=justificacion_memoria,
                    es_contratacion=True,
                    estado=MemoriaCalculo.EstadoMemoria.APROBADO_FINANZAS,
                    fecha_aprobacion=timezone.now()
                )
                count_memorias += 1

                # Crear Detalles
                for item in items:
                    DetallePresupuestoMemoria.objects.create(
                        memoria=memoria,
                        partida=partida,
                        descripcion=item['descripcion'],
                        unidad_medida=item['unidad_medida'],
                        fuente_excel=sname_clean[:120],
                        factor_calculo=item['factor_calculo'],
                        total_programado=item['total_origen'],
                        cantidad=item['cantidad'],
                        precio_unitario=item['precio_unitario'],
                        estado_ejecucion=DetallePresupuestoMemoria.EstadoGasto.PENDIENTE
                    )
                    count_detalles += 1

                recalcular_saldos_memoria(memoria)

                # Acumular presupuesto del área
                area_totals[area.id] = area_totals.get(area.id, Decimal('0.00')) + sheet_total
                total_presupuesto_2026 += sheet_total

            # Guardar presupuestos consolidados por Área para la gestión 2026
            for area_id, monto_area in area_totals.items():
                presupuesto_area, _ = PresupuestoArea.objects.get_or_create(
                    gestion=gestion,
                    area_id=area_id,
                    defaults={'monto_inicial': Decimal('0.00'), 'monto_actual': Decimal('0.00')}
                )
                presupuesto_area.monto_inicial = monto_area
                presupuesto_area.monto_actual = monto_area
                presupuesto_area.save(update_fields=['monto_inicial', 'monto_actual', 'updated_at'])

            self.stdout.write(self.style.SUCCESS("=" * 60))
            self.stdout.write(self.style.SUCCESS(
                f"[OK] Carga POA 2026 completada exitosamente!\n"
                f"   - Memorias válidas creadas    : {count_memorias}\n"
                f"   - Hojas con TOTAL 0,00 omitidas: {count_omitidas_cero}\n"
                f"   - Detalles presupuestarios    : {count_detalles}\n"
                f"   - Presupuesto Total 2026      : Bs {total_presupuesto_2026:,.2f}\n"
                f"   - Estado de memorias          : APROBADO_FINANZAS"
            ))
            self.stdout.write(self.style.SUCCESS("=" * 60))
