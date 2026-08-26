"""Importación verificable de memorias de cálculo POA 2027 por área.

El comando no modifica la base de datos salvo que se indique --apply.  La
simulación valida las hojas, sus totales y la correspondencia contra la hoja
de determinación de requerimientos antes de permitir una carga real.
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.memorias.models import DetallePresupuestoMemoria, MemoriaCalculo
from apps.memorias.utils import recalcular_saldos_memoria
from apps.organizacional.models import Area, Programa, Seccion
from apps.planificacion.models import AccionCortoPlazo, AccionMedianoPlazo, Operacion
from apps.presupuestos.models import Gestion, Partida, PresupuestoArea


@dataclass(frozen=True)
class AreaConfig:
    nombre: str
    programa: str
    aliases_requerimientos: tuple[str, ...]


# Siglas de las hojas de memorias y su nombre institucional.  Las siglas CIJ,
# GYA, RIB y TDD se conservan para mantener trazabilidad con el archivo fuente.
AREAS = {
    'PL': AreaConfig('Unidad de Planificación', '1', ('UNIDAD DE PLANIFICACION',)),
    'UT': AreaConfig('Unidad de Transparencia', '1', ('UNIDAD DE TRANSPARENCIA',)),
    'AI': AreaConfig('Unidad de Auditoría Interna', '1', ('UNIDAD DE AUDITORIA INTERNA', 'UAI')),
    'AJ': AreaConfig('Unidad Jurídica', '1', ('UNIDAD JURIDICA', 'UNIDAD DE JURIDICA', 'UNIDAD DE ASESORIA JURIDICA')),
    'CIAC': AreaConfig('CIAC', '1', ('CIAC', 'CIAC TAMEP')),
    'OD': AreaConfig('ODECO', '1', ('ODECO',)),
    'IF': AreaConfig('Gerencia de Informática', '1', ('GERENCIA DE INFORMATICA',)),
    'GAA': AreaConfig('Gerencia de Asuntos Administrativos', '2', ('GERENCIA DE ASUNTOS ADMINISTRATIVOS',)),
    'GC': AreaConfig('Gerencia Comercial', '410', ('GERENCIA COMERCIAL',)),
    'GO': AreaConfig('Gerencia de Operaciones', '210', ('GERENCIA DE OPERACIONES',)),
    'AE': AreaConfig('Gerencia de Aeronavegabilidad', '210', ('GERENCIA DE AERONAVEGABILIDAD',)),
    'SMS': AreaConfig('Gerencia de SMS (AVSEC)', '210', ('GERENCIA SMS', 'GERENCIA DE SMS', 'SMS', 'GERENCIA SEGURIDAD OPERACIONAL Y ASEGURAMIENTO A LA CALIDAD')),
    'EA': AreaConfig('Gerencia Regional El Alto', '410', ('GERENCIA REGIONAL EL ALTO',)),
    'LP': AreaConfig('Sucursal La Paz (Agencia Montes)', '410', ('SUCURSAL LA PAZ', 'AGENCIA MONTES')),
    'CB': AreaConfig('Gerencia Regional Cochabamba', '410', ('GERENCIA REGIONAL COCHABAMBA',)),
    'SC': AreaConfig('Gerencia Regional Santa Cruz', '410', ('GERENCIA REGIONAL SANTA CRUZ',)),
    'CIJ': AreaConfig('Gerencia Regional Cobija', '410', ('GERENCIA REGIONAL COBIJA',)),
    'GYA': AreaConfig('Agencia Guayaramerín', '410', ('AGENCIA GUAYARAMERIN', 'GUAYARAMERIN')),
    'RIB': AreaConfig('Agencia Riberalta', '410', ('AGENCIA RIBERALTA', 'RIBERALTA')),
    'TDD': AreaConfig('Agencia Trinidad', '410', ('AGENCIA TRINIDAD', 'TRINIDAD')),
}


def text(value) -> str:
    return str(value or '').strip()


def key(value) -> str:
    normalized = unicodedata.normalize('NFKD', text(value)).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'[^A-Z0-9]+', ' ', normalized.upper()).strip()


def decimal(value, default=Decimal('0')) -> Decimal:
    if value is None or text(value) == '':
        return default
    try:
        return Decimal(str(value).replace(',', '.')).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        return default


class Command(BaseCommand):
    help = 'Simula o importa las memorias POA 2027 de una sola área desde los dos Excel oficiales.'

    def add_arguments(self, parser):
        parser.add_argument('--area', required=True, choices=sorted(AREAS), help='Sigla del área a procesar, por ejemplo PL.')
        parser.add_argument('--consolidado', required=True, help='Ruta al archivo “27.7 consolidado general 2027”.')
        parser.add_argument('--operaciones', required=True, help='Ruta al archivo “OPERACION-PRESUPUESTO 2027”.')
        parser.add_argument('--apply', action='store_true', help='Escribe los datos en la base de datos. Sin esta opción solo simula.')
        parser.add_argument('--prioridad-consolidado', action='store_true', help='Permite importar una memoria cuyo monto no esté correlacionado en Operación-Presupuesto; conserva el consolidado como fuente autoritativa y deja la correlación explícitamente pendiente.')
        parser.add_argument('--report', help='Ruta opcional donde se guardará el reporte JSON de la ejecución.')

    def handle(self, *args, **options):
        area_code = options['area']
        consolidado_path = Path(options['consolidado']).expanduser().resolve()
        operaciones_path = Path(options['operaciones']).expanduser().resolve()
        for path in (consolidado_path, operaciones_path):
            if not path.is_file():
                raise CommandError(f'No se encontró el archivo: {path}')

        self.stdout.write(f"{'IMPORTACIÓN' if options['apply'] else 'SIMULACIÓN'} POA 2027 — área {area_code}")
        requirements = self._read_requirements(operaciones_path)
        memories = self._read_memories(consolidado_path, area_code)
        report = self._validate(area_code, memories, requirements, options['prioridad_consolidado'])

        self._print_report(report)
        if report['errors']:
            self._save_report(options.get('report'), report)
            raise CommandError('La carga fue bloqueada: corrija las inconsistencias reportadas antes de usar --apply.')

        if options['apply']:
            with transaction.atomic():
                self._persist(area_code, memories, requirements, options['prioridad_consolidado'])
            report['applied'] = True
            self.stdout.write(self.style.SUCCESS('Importación completada y confirmada en una transacción.'))
        else:
            self.stdout.write(self.style.WARNING('No se escribió ningún dato. Revise el reporte y repita con --apply cuando sea correcto.'))
        self._save_report(options.get('report'), report)

    def _read_requirements(self, path: Path):
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        records = []
        for sheet_name in ('PROG 1', 'PROG 2', 'PROG 410', 'PROG 210'):
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            current = {'codigo': '', 'accion': '', 'area': '', 'operacion': ''}
            for row in sheet.iter_rows(min_row=5, values_only=True):
                values = list(row) + [None] * 11
                if values[3] is not None:
                    current['codigo'] = text(values[3])
                if values[4] is not None:
                    current['accion'] = text(values[4])
                if values[5] is not None:
                    current['area'] = text(values[5])
                if values[6] is not None:
                    current['operacion'] = text(values[6])
                partida = text(values[7])
                if not re.fullmatch(r'\d{5}', partida):
                    continue
                records.append({
                    **current,
                    'programa': sheet_name.replace('PROG ', ''),
                    'partida': partida,
                    'bien_servicio': text(values[8]),
                    'mes_requerido': text(values[9]),
                    'presupuesto': decimal(values[10]),
                })
        return records

    def _read_memories(self, path: Path, area_code: str):
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        selected = [name for name in workbook.sheetnames if re.match(rf'^\s*\d{{5}}\s*{area_code}\s*$', name, re.I)]
        if not selected:
            raise CommandError(f'No se encontraron hojas de memoria para el área {area_code}.')
        memories = []
        for sheet_name in selected:
            sheet = workbook[sheet_name]
            header_row, columns = self._find_memory_header(sheet)
            partida_match = re.search(r'\d{5}', sheet_name)
            partida = partida_match.group(0)
            items, source_total, last_description = [], None, ''
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                values = list(row)
                description = text(values[columns['descripcion']]) if len(values) > columns['descripcion'] else ''
                number = text(values[columns['numero']]) if len(values) > columns['numero'] else ''
                if 'TOTAL' in key(number) or 'TOTAL' in key(description):
                    if len(values) > columns['total']:
                        source_total = decimal(values[columns['total']])
                    break
                if not description:
                    quantity_candidate = decimal(values[columns['cantidad']] if len(values) > columns['cantidad'] else None)
                    price_candidate = decimal(values[columns['precio']] if len(values) > columns['precio'] else None)
                    # Una fila puede continuar un mismo requerimiento (por
                    # ejemplo, otro tipo de hora de simulador) sin repetir la
                    # descripción. Se conserva la descripción anterior para
                    # no omitir un importe presupuestado.
                    if last_description and quantity_candidate and price_candidate:
                        description = f'{last_description} (continuación)'
                    else:
                    # Algunas hojas (por ejemplo 34400OD) omiten la etiqueta
                    # “TOTAL PARTIDA”, pero dejan el importe total en la
                    # columna TOTAL inmediatamente después de los ítems.
                        candidate_total = decimal(values[columns['total']]) if len(values) > columns['total'] else Decimal('0')
                        if items and candidate_total and (number in {'.', '-'} or not number):
                            source_total = candidate_total
                            break
                        continue
                quantity = decimal(values[columns['cantidad']] if len(values) > columns['cantidad'] else None, Decimal('1'))
                unit = text(values[columns['unidad']] if len(values) > columns['unidad'] else '') or 'UNIDAD'
                price = decimal(values[columns['precio']] if len(values) > columns['precio'] else None)
                line_total = decimal(values[columns['total']] if len(values) > columns['total'] else None)
                base_total = quantity * price
                factor = (line_total / base_total).quantize(Decimal('0.0001')) if base_total else Decimal('1.0000')
                justification = text(values[columns['justificacion']] if len(values) > columns['justificacion'] else '')
                items.append({'descripcion': description, 'cantidad': quantity, 'unidad_medida': unit, 'precio_unitario': price, 'factor_calculo': factor, 'total_origen': line_total, 'justificacion': justification})
                last_description = description
            if not items:
                raise CommandError(f'La hoja {sheet_name} no contiene detalles importables.')
            calculated_total = sum((item['total_origen'] for item in items), Decimal('0')).quantize(Decimal('0.01'))
            memories.append({'hoja': sheet_name, 'partida': partida, 'items': items, 'total_origen': source_total, 'total_calculado': calculated_total})
        return memories

    def _find_memory_header(self, sheet):
        columns = {'numero': 0, 'descripcion': 1, 'cantidad': 3, 'unidad': 4, 'precio': 5, 'total': 6, 'justificacion': 7}
        for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            labels = [key(value) for value in row]
            if not any('DESCRIPCI' in label for label in labels):
                continue
            for index, label in enumerate(labels):
                if label in {'N', 'NO', 'NRO', 'NUMERO'}:
                    columns['numero'] = index
                elif 'DESCRIPCI' in label:
                    columns['descripcion'] = index
                elif 'CANTIDAD' in label:
                    columns['cantidad'] = index
                elif 'UNIDAD' in label:
                    columns['unidad'] = index
                elif 'PRECIO' in label or 'UNITARIO' in label:
                    columns['precio'] = index
                elif label == 'TOTAL':
                    columns['total'] = index
                elif 'JUSTIFICACI' in label:
                    columns['justificacion'] = index
            return row_number, columns
        raise CommandError(f'No se encontró la cabecera de detalle en la hoja {sheet.title}.')

    def _requirements_for_area(self, area_code, requirements):
        aliases = {key(alias) for alias in AREAS[area_code].aliases_requerimientos}
        return [record for record in requirements if key(record['area']) in aliases]

    def _validate(self, area_code, memories, requirements, prioridad_consolidado=False):
        area_requirements = self._requirements_for_area(area_code, requirements)
        errors, warnings, rows = [], [], []
        if not area_requirements and not prioridad_consolidado:
            errors.append(f'No se hallaron requerimientos para el área {area_code} en OPERACION-PRESUPUESTO 2027.')
        for memory in memories:
            if memory['total_origen'] is None:
                errors.append(f"{memory['hoja']}: no se identificó el total de la partida.")
            elif memory['total_calculado'] != memory['total_origen']:
                errors.append(f"{memory['hoja']}: total calculado {memory['total_calculado']} != total del Excel {memory['total_origen']}.")
            candidates = [r for r in area_requirements if r['partida'] == memory['partida'] and r['presupuesto'] == memory['total_calculado']]
            if len(candidates) != 1:
                if prioridad_consolidado and not candidates:
                    warnings.append(f"{memory['hoja']}: se usará el consolidado como fuente autoritativa; no existe correlación de operación/mes en OPERACION-PRESUPUESTO.")
                    rows.append({'hoja': memory['hoja'], 'partida': memory['partida'], 'items': len(memory['items']), 'total': str(memory['total_calculado']), 'operacion': 'SIN CORRELACIÓN EN OPERACIÓN-PRESUPUESTO', 'mes_requerido': ''})
                    continue
                errors.append(f"{memory['hoja']}: se esperó 1 requerimiento con partida {memory['partida']} y Bs {memory['total_calculado']}; se encontraron {len(candidates)}.")
                continue
            requirement = candidates[0]
            rows.append({'hoja': memory['hoja'], 'partida': memory['partida'], 'items': len(memory['items']), 'total': str(memory['total_calculado']), 'operacion': requirement['operacion'], 'mes_requerido': requirement['mes_requerido']})
        return {'area': area_code, 'memorias': len(memories), 'items': sum(len(m['items']) for m in memories), 'total': str(sum((m['total_calculado'] for m in memories), Decimal('0'))), 'rows': rows, 'errors': errors, 'warnings': warnings, 'applied': False}

    def _persist(self, area_code, memories, requirements, prioridad_consolidado=False):
        config = AREAS[area_code]
        program, _ = Programa.objects.get_or_create(codigo=config.programa, defaults={'nombre': f'Programa {config.programa}'})
        area, _ = Area.objects.get_or_create(programa=program, nombre=config.nombre, defaults={'codigo': f'P-{config.programa}-{area_code}', 'tipo': Area.TipoArea.UNIDAD if area_code in {'PL', 'UT', 'AI', 'AJ', 'CIAC', 'OD'} else Area.TipoArea.GERENCIA})
        section, _ = Seccion.objects.get_or_create(area=area, nombre=config.nombre)
        gestion, _ = Gestion.objects.get_or_create(anio=2027, defaults={'estado': Gestion.EstadoGestion.FORMULACION})
        area_requirements = self._requirements_for_area(area_code, requirements)
        total_area = Decimal('0')
        for memory_data in memories:
            candidates = [r for r in area_requirements if r['partida'] == memory_data['partida'] and r['presupuesto'] == memory_data['total_calculado']]
            if candidates:
                requirement = candidates[0]
            elif prioridad_consolidado:
                requirement = {
                    'programa': config.programa,
                    'codigo': config.programa,
                    'accion': f'Requerimientos POA 2027 del Programa {config.programa}.',
                    'area': config.nombre,
                    'operacion': 'REQUERIMIENTO CONSOLIDADO SIN CORRELACION EN OPERACION-PRESUPUESTO',
                    'partida': memory_data['partida'],
                    'bien_servicio': '',
                    'mes_requerido': '',
                    'presupuesto': memory_data['total_calculado'],
                }
            else:
                raise CommandError(f"No existe requerimiento correlacionado para {memory_data['hoja']}.")
            operation = self._get_operation(gestion, program, area, requirement)
            partida, _ = Partida.objects.get_or_create(codigo=memory_data['partida'], clase=Partida.ClasePartida.EGRESO, defaults={'nombre': requirement['bien_servicio'] or f"Partida {memory_data['partida']}"})
            sheet_code = re.sub(r'\s+', '', memory_data['hoja'])
            code = f"MEM-2027-{sheet_code}"
            if MemoriaCalculo.objects.filter(codigo=code).exists():
                raise CommandError(f'Ya existe la memoria {code}; la importación fue cancelada para evitar duplicados.')
            justification = next((item['justificacion'] for item in memory_data['items'] if item['justificacion']), f"Importado de la hoja {memory_data['hoja']}.")
            memory = MemoriaCalculo.objects.create(codigo=code, gestion=gestion, seccion=section, operacion=operation, justificacion=justification, es_contratacion=True, estado=MemoriaCalculo.EstadoMemoria.BORRADOR)
            for item in memory_data['items']:
                DetallePresupuestoMemoria.objects.create(memoria=memory, partida=partida, descripcion=item['descripcion'], unidad_medida=item['unidad_medida'][:50], mes_requerido=requirement['mes_requerido'][:120], fuente_excel=memory_data['hoja'][:120], factor_calculo=item['factor_calculo'], cantidad=item['cantidad'], precio_unitario=item['precio_unitario'])
            recalcular_saldos_memoria(memory)
            total_area += memory_data['total_calculado']
        budget, _ = PresupuestoArea.objects.get_or_create(gestion=gestion, area=area, defaults={'monto_inicial': Decimal('0'), 'monto_actual': Decimal('0')})
        budget.monto_inicial += total_area
        budget.monto_actual += total_area
        budget.save(update_fields=['monto_inicial', 'monto_actual', 'updated_at'])

    def _get_operation(self, gestion, program, area, requirement):
        amp, _ = AccionMedianoPlazo.objects.get_or_create(codigo=f'AMP-2027-P{program.codigo}', defaults={'programa': program, 'descripcion': f'Planificación estratégica institucional del Programa {program.codigo} para la gestión 2027.', 'periodo_inicio': 2026, 'periodo_fin': 2030})
        acp, _ = AccionCortoPlazo.objects.get_or_create(codigo=f'ACP-2027-P{program.codigo}', defaults={'accion_mediano_plazo': amp, 'gestion': gestion, 'descripcion': requirement['accion']})
        operation_key = re.sub(r'[^A-Z0-9]', '', key(requirement['operacion']))[:12]
        code = f'OP-2027-{area.codigo[-4:]}-{operation_key}'[:30]
        operation, _ = Operacion.objects.get_or_create(codigo=code, defaults={'accion_corto_plazo': acp, 'area': area, 'descripcion': requirement['operacion'], 'es_contratacion': True})
        return operation

    def _print_report(self, report):
        self.stdout.write(f"Memorias: {report['memorias']} | Detalles: {report['items']} | Total: Bs {report['total']}")
        for row in report['rows']:
            self.stdout.write(f"  OK {row['hoja']}: {row['items']} ítem(s), Bs {row['total']} — {row['mes_requerido']}")
        for warning in report['warnings']:
            self.stdout.write(self.style.WARNING(f'ADVERTENCIA: {warning}'))
        for error in report['errors']:
            self.stdout.write(self.style.ERROR(f'ERROR: {error}'))

    def _save_report(self, report_path, report):
        if not report_path:
            return
        path = Path(report_path).expanduser().resolve()
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
        self.stdout.write(f'Reporte guardado en: {path}')
